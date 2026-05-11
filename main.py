import os
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl
import psycopg2

COLOR_MAIN = "#FFFFFF"    # Белый фон
COLOR_SEC = "#BBD9B2"     # Светло-зеленый (меню)
COLOR_ACCENT = "#2D6033"  # Темно-зеленый (кнопки, акценты)
FONT_NAME = "Gabriola"
F_SIZE = 14

# Пути к ресурсам и данные для подключения к базе
BASE_PATH = r"A:\programming\python\demo\Прил_В2_КОД 09.02.07-2-2025-БУ\Ресурсы"
ICON_FILE = r"A:\programming\python\demo\Прил_В2_КОД 09.02.07-2-2025-БУ\Ресурсы\Наш декор.ico"

DB_CONFIG = {
    "host": "192.168.0.4",
    "port": 5432,
    "database": "demo",
    "user": "postgres",
    "password": "newroot"
}

# Список типов продукции для выпадающего меню
PRODUCT_TYPES = ["Декоративные обои", "Фотообои", "Обои под покраску", "Стеклообои"]

class DecorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Производственная компания «Наш декор»")
        self.root.geometry("1200x900")
        self.root.configure(bg=COLOR_MAIN)

        # Ставим иконку в верхний угол окна
        self.set_app_icon()

        # Подключаемся к базе и подготавливаем данные
        try:
            self.conn = psycopg2.connect(**DB_CONFIG)
            self.setup_database_structure()
            self.import_data_from_excel()
        except Exception as e:
            messagebox.showerror("Ошибка БД", f"Не удалось подключиться к PostgreSQL:\n{e}")
            self.conn = None

        self.setup_ui()

    def set_app_icon(self):
        """Установка иконки окна из .ico файла"""
        try:
            # Для .ico файлов используется метод iconbitmap
            self.root.iconbitmap(ICON_FILE)
        except Exception as e:
            print(f"Не удалось установить иконку: {e}")


    def setup_database_structure(self):
        """Создаем все таблицы в БД. Используем DROP, чтобы структура всегда была актуальной"""
        with self.conn.cursor() as cur:
            cur.execute("DROP TABLE IF EXISTS product_materials CASCADE;")
            cur.execute("DROP TABLE IF EXISTS products CASCADE;")
            cur.execute("DROP TABLE IF EXISTS materials CASCADE;")
            cur.execute("DROP TABLE IF EXISTS coefficients CASCADE;")
            cur.execute("DROP TABLE IF EXISTS scrap CASCADE;")

            cur.execute("CREATE TABLE products (article TEXT PRIMARY KEY, name TEXT NOT NULL, type TEXT, min_price REAL, width REAL)")
            cur.execute("CREATE TABLE materials (name TEXT PRIMARY KEY, type TEXT, price REAL, stock REAL, min_stock REAL, pkg REAL, unit TEXT)")
            cur.execute("CREATE TABLE coefficients (type TEXT PRIMARY KEY, val REAL)")
            cur.execute("CREATE TABLE scrap (type TEXT PRIMARY KEY, percent REAL)")
            cur.execute("CREATE TABLE product_materials (product_name TEXT, material_name TEXT, quantity REAL, PRIMARY KEY (product_name, material_name))")
            self.conn.commit()

    def import_data_from_excel(self):
        """Читаем все .xlsx файлы из папки Ресурсы и заливаем их в БД"""
        if not os.path.exists(BASE_PATH): return

        files = [f for f in os.listdir(BASE_PATH) if f.endswith('.xlsx')]
        with self.conn.cursor() as cur:
            for file in files:
                full_path = os.path.join(BASE_PATH, file)
                wb = openpyxl.load_workbook(full_path, data_only=True)
                sheet = wb.active
                first_cell = sheet.cell(row=1, column=1).value

                if first_cell == "Продукция":
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0] and row[1]:
                            cur.execute("INSERT INTO product_materials VALUES (%s, %s, %s)",
                                        (row[0], row[1], float(row[2]) if row[2] else 0))
                elif first_cell == "Тип материала":
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            val = row[1]
                            if isinstance(val, str): val = float(val.replace('%', '').replace(',', '.'))
                            cur.execute("INSERT INTO scrap VALUES (%s, %s)", (row[0], val))
                elif first_cell == "Тип продукции" and sheet.cell(row=1, column=3).value == "Артикул":
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[2]:
                            cur.execute("INSERT INTO products VALUES (%s, %s, %s, %s, %s)",
                                        (str(row[2]), row[1], row[0], float(row[3]), float(row[4])))
                elif first_cell == "Тип продукции" and sheet.cell(row=1, column=2).value == "Коэффициент типа продукции":
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]: cur.execute("INSERT INTO coefficients VALUES (%s, %s)", (row[0], float(row[1])))
                elif first_cell == "Наименование материала":
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if row[0]:
                            cur.execute("INSERT INTO materials VALUES (%s, %s, %s, %s, %s, %s, %s)",
                                        (row[0], row[1], float(row[2]), float(row[3]), float(row[4]), float(row[5]), row[6]))
            self.conn.commit()

    def setup_ui(self):
        try:
            logo_path = os.path.join(BASE_PATH, ICON_FILE)
            self.logo_img = tk.PhotoImage(file=logo_path)
            tk.Label(self.root, image=self.logo_img, bg=COLOR_MAIN).pack(pady=20)
        except:
            tk.Label(self.root, text="НАШ ДЕКОР", font=(FONT_NAME, 24, "bold"), fg=COLOR_ACCENT, bg=COLOR_MAIN).pack(pady=20)

        self.main_frame = tk.Frame(self.root, bg=COLOR_MAIN)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.nav_bar = tk.Frame(self.main_frame, bg=COLOR_SEC, width=200)
        self.nav_bar.pack(side=tk.LEFT, fill=tk.Y)

        menu = [("Продукция", self.show_products), ("Склад материалов", self.show_materials)]
        for text, cmd in menu:
            btn = tk.Button(self.nav_bar, text=text, command=cmd, bg=COLOR_ACCENT, fg="white",
                            font=(FONT_NAME, F_SIZE), relief=tk.FLAT, pady=10)
            btn.pack(fill=tk.X, padx=10, pady=5)

        self.content = tk.Frame(self.main_frame, bg=COLOR_MAIN)
        self.content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20)
        self.show_products()

    def clear(self):
        for w in self.content.winfo_children(): w.destroy()

    def show_products(self):
        self.clear()
        tk.Label(self.content, text="Каталог продукции", font=(FONT_NAME, 20), fg=COLOR_ACCENT, bg=COLOR_MAIN).pack(pady=10)

        tk.Button(self.content, text="+ Добавить новый товар", command=self.open_product_form,
                  bg=COLOR_ACCENT, fg="white", font=(FONT_NAME, 12), pady=5).pack(pady=10)

        self.canvas = tk.Canvas(self.content, bg=COLOR_MAIN, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.content, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=COLOR_MAIN)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        query = "SELECT p.article, p.name, p.type, (p.min_price * COALESCE(c.val, 1)) as calc_price, p.width FROM products p LEFT JOIN coefficients c ON p.type = c.type"
        with self.conn.cursor() as cur:
            cur.execute(query)
            for row in cur.fetchall():
                self.create_product_card(row)

    def create_product_card(self, data):
        article, name, p_type, price, width = data
        card = tk.Frame(self.scrollable_frame, bg=COLOR_MAIN, highlightbackground="grey", highlightthickness=1, padx=15, pady=10)
        card.pack(fill=tk.X, padx=20, pady=10)

        top_row = tk.Frame(card, bg=COLOR_MAIN)
        top_row.pack(fill=tk.X)
        tk.Label(top_row, text=f"{p_type} | {name}", font=(FONT_NAME, 16, "bold"), bg=COLOR_MAIN, fg=COLOR_ACCENT).pack(side=tk.LEFT)
        tk.Label(top_row, text=f"Стоимость: {price:.2f} р.", font=(FONT_NAME, 16), bg=COLOR_MAIN).pack(side=tk.RIGHT)

        tk.Label(card, text=f"Артикул: {article}", font=(FONT_NAME, 12), bg=COLOR_MAIN).pack(anchor="w")
        tk.Label(card, text=f"Ширина: {width} м", font=(FONT_NAME, 12), bg=COLOR_MAIN).pack(anchor="w")

        btn_row = tk.Frame(card, bg=COLOR_MAIN)
        btn_row.pack(fill=tk.X, pady=5)

        def edit_this():
            with self.conn.cursor() as cur:
                cur.execute("SELECT article, name, type, min_price, width FROM products WHERE article=%s", (article,))
                self.open_product_form(cur.fetchone())

        tk.Button(btn_row, text="Редактировать", command=edit_this, bg=COLOR_ACCENT, fg="white", font=(FONT_NAME, 10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_row, text="Состав материалов", command=lambda n=name: self.show_product_composition(n), bg=COLOR_ACCENT, fg="white", font=(FONT_NAME, 10)).pack(side=tk.LEFT, padx=5)

    def show_product_composition(self, product_name):
        win = tk.Toplevel(self.root)
        win.title(f"Состав: {product_name}")
        win.geometry("500x400")
        win.configure(bg=COLOR_MAIN)
        tk.Label(win, text=f"Материалы для: {product_name}", font=(FONT_NAME, 16), bg=COLOR_MAIN, fg=COLOR_ACCENT).pack(pady=10)
        tree = ttk.Treeview(win, columns=("Mat", "Qty"), show="headings")
        tree.heading("Mat", text="Наименование материала")
        tree.heading("Qty", text="Кол-во")
        with self.conn.cursor() as cur:
            cur.execute("SELECT material_name, quantity FROM product_materials WHERE product_name = %s", (product_name,))
            for row in cur.fetchall(): tree.insert("", tk.END, values=row)
        tree.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    def open_product_form(self, edit_data=None):
        win = tk.Toplevel(self.root)
        win.title("Карточка товара")
        win.geometry("400x500")
        win.configure(bg=COLOR_MAIN)

        field_info = [("Артикул", "art"), ("Наименование", "name"), ("Тип", "type"), ("Мин. Цена", "price"), ("Ширина", "width")]
        entries = {}

        for label_text, key in field_info:
            tk.Label(win, text=label_text, bg=COLOR_MAIN, font=(FONT_NAME, 12)).pack(pady=5)
            if key == "type":
                combo = ttk.Combobox(win, values=PRODUCT_TYPES, state="readonly")
                combo.pack(pady=5)
                entries[key] = combo
            else:
                entry = tk.Entry(win)
                entry.pack(pady=5)
                entries[key] = entry

        if edit_data:
            entries['art'].insert(0, edit_data[0])
            entries['name'].insert(0, edit_data[1])
            entries['type'].set(edit_data[2])
            entries['price'].insert(0, edit_data[3])
            entries['width'].insert(0, edit_data[4])
            entries['art'].config(state='disabled')

        def save():
            try:
                art = entries['art'].get().strip()
                name = entries['name'].get().strip()
                p_type = entries['type'].get().strip()
                price_str = entries['price'].get().strip()
                width_str = entries['width'].get().strip()

                if not all([art, name, p_type, price_str, width_str]):
                    raise ValueError("Все поля должны быть заполнены!")

                try:
                    price = float(price_str.replace(',', '.'))
                    width = float(width_str.replace(',', '.'))
                except ValueError:
                    raise ValueError("Цена и ширина должны быть числами!")

                if price < 0 or width < 0:
                    raise ValueError("Цена и ширина не могут быть отрицательными!")

                with self.conn.cursor() as cur:
                    cur.execute("""
                                INSERT INTO products (article, name, type, min_price, width)
                                VALUES (%s, %s, %s, %s, %s)
                                ON CONFLICT (article) DO UPDATE SET name=%s, type=%s, min_price=%s, width=%s
                                """, (art, name, p_type, price, width, name, p_type, price, width))
                    self.conn.commit()

                messagebox.showinfo("Успех", "Данные успешно сохранены!")
                win.destroy()
                self.show_products()
            except Exception as e:
                messagebox.showerror("Ошибка ввода", str(e))

        tk.Button(win, text="Сохранить", command=save, bg=COLOR_ACCENT, fg="white", font=(FONT_NAME, 12), pady=10).pack(pady=20)

    def show_materials(self):
        """Просмотр списка всех материалов со склада"""
        self.clear()
        tk.Label(self.content, text="Склад материалов", font=(FONT_NAME, 20), fg=COLOR_ACCENT, bg=COLOR_MAIN).pack(pady=10)

        # Определяем список колонок один раз
        cols_ids = ("C0", "C1", "C2", "C3", "C4", "C5", "C6")
        cols_text = ["Наименование", "Тип", "Цена", "Остаток", "Мин.", "Упак.", "Ед."]

        tree = ttk.Treeview(self.content, columns=cols_ids, show="headings")

        # Теперь привязываем текст к ID колонок
        for i in range(len(cols_ids)):
            tree.heading(cols_ids[i], text=cols_text[i])

        with self.conn.cursor() as cur:
            cur.execute("SELECT * FROM materials")
            for row in cur.fetchall(): tree.insert("", tk.END, values=row)
        tree.pack(fill=tk.BOTH, expand=True)

if __name__ == "__main__":
    root = tk.Tk()
    app = DecorApp(root)
    root.mainloop()
