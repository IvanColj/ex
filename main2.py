import math
import os
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl
import psycopg2

C_MAIN = "#FFFFFF"    # Основной фон (Белый)
C_SEC = "#ABCFCE"     # Дополнительный фон (Светло-голубой)
C_ACCENT = "#546F94"  # Акцент для кнопок и заголовков (Синий)
FONT_NAME = "Comic Sans MS"
F_SIZE = 12

# Пути к ресурсам (Иконка и Логотип)
BASE_PATH = r"A:\programming\python\Демонстрационный экзамен_лето_2025\Задание_2996\Прил_В1_КОД 09.02.07-2-2025-БУ\Ресурсы"
LOGO_FILE = "logo.png"
ICON_FILE = r"A:\programming\python\Демонстрационный экзамен_лето_2025\Задание_2996\Прил_В1_КОД 09.02.07-2-2025-БУ\Ресурсы\Мозаика.ico"

# Реквизиты подключения к PostgreSQL
DB_CONFIG = {
    "host": "192.168.0.4",
    "port": 5432,
    "database": "demo2",
    "user": "postgres",
    "password": "newroot"
}

class MaterialApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Управление материалами (demo2)")
        self.root.geometry("1200x900")
        self.root.configure(bg=C_MAIN)

        # Устанавливаем иконку окна .ico (Критерий Г1Д1)
        self.set_app_icon()

        try:
            # Пытаемся подключиться к БД. Если падает здесь -> приложение закрывается
            self.conn = psycopg2.connect(**DB_CONFIG)
            self.setup_database_structure() # Создаем таблицы
            self.import_data_from_excel()   # Загружаем данные из файлов
        except Exception as e:
            messagebox.showerror("Критическая ошибка", f"База данных недоступна:\n{e}")
            self.conn = None

        # Если соединение с БД успешно, рисуем интерфейс
        if self.conn:
            self.setup_ui()
        else:
            self.root.destroy()

    def set_app_icon(self):
        """Установка иконки в заголовке окна Windows"""
        try:
            self.root.iconbitmap(ICON_FILE)
        except:
            pass # Если файл иконки не найден, программа продолжит работу

    def setup_database_structure(self):
        """
        Инициализация структуры БД. 
        Используем CASCADE, чтобы корректно удалять таблицы со связями.
        """
        with self.conn.cursor() as cur:
            cur.execute("DROP TABLE IF EXISTS materials CASCADE;")
            cur.execute("DROP TABLE IF EXISTS material_types CASCADE;")
            cur.execute("DROP TABLE IF EXISTS suppliers CASCADE;")
            cur.execute("DROP TABLE IF EXISTS scrap CASCADE;")
            cur.execute("DROP TABLE IF EXISTS coefficients CASCADE;")
            cur.execute("DROP TABLE IF EXISTS products CASCADE;")

            # 1. Справочник типов (для выпадающего списка в форме)
            cur.execute("CREATE TABLE material_types (id SERIAL PRIMARY KEY, name TEXT UNIQUE NOT NULL);")

            # 2. Основная таблица материалов (Связь с material_types через внешний ключ)
            cur.execute("""
                        CREATE TABLE materials (
                                                   id SERIAL PRIMARY KEY,
                                                   name TEXT NOT NULL,
                                                   type_id INTEGER REFERENCES material_types(id),
                                                   price REAL NOT NULL,
                                                   stock REAL NOT NULL,
                                                   min_qty REAL NOT NULL,
                                                   pack_size REAL NOT NULL,
                                                   unit TEXT NOT NULL
                        )
                        """)

            # 3. Поставщики (Наименование как первичный ключ для уникальности)
            cur.execute("CREATE TABLE suppliers (name TEXT PRIMARY KEY, type TEXT, inn TEXT, rating INTEGER, start_date DATE);")

            # 4. Таблицы для вспомогательных данных (Брак, Коэффициенты, Продукция)
            cur.execute("CREATE TABLE scrap (type TEXT PRIMARY KEY, percent REAL);")
            cur.execute("CREATE TABLE coefficients (type TEXT PRIMARY KEY, val REAL);")
            cur.execute("CREATE TABLE products (article TEXT PRIMARY KEY, name TEXT, type TEXT, price REAL, width REAL);")

            # Заполняем базовые типы материалов из ТЗ, чтобы импорт из Excel прошел успешно
            types = ('Пластичные материалы', 'Добавка', 'Электролит', 'Глазурь', 'Пигмент')
            for t in types:
                cur.execute("INSERT INTO material_types (name) VALUES (%s)", (t,))
            self.conn.commit()

    def import_data_from_excel(self):
        """
        Универсальный импорт всех .xlsx файлов из папки ресурсов.
        Добавлена строгая проверка длины строки (len(row)), чтобы избежать ошибки 'tuple index out of range'.
        """
        if not os.path.exists(BASE_PATH): return
        files = [f for f in os.listdir(BASE_PATH) if f.endswith('.xlsx')]

        with self.conn.cursor() as cur:
            for file in files:
                try:
                    wb = openpyxl.load_workbook(os.path.join(BASE_PATH, file), data_only=True)
                    sheet = wb.active
                    first_cell = sheet.cell(row=1, column=1).value

                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        if not row or not row[0]: continue # Пропускаем пустые строки

                        # ИМПОРТ МАТЕРИАЛОВ (ждем 7 колонок)
                        if first_cell == "Наименование материала" and len(row) >= 7:
                            cur.execute("SELECT id FROM material_types WHERE name = %s", (row[1],))
                            res = cur.fetchone()
                            t_id = res[0] if res else None
                            cur.execute("INSERT INTO materials (name, type_id, price, stock, min_qty, pack_size, unit) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                                        (row[0], t_id, float(row[2]), float(row[3]), float(row[4]), float(row[5]), row[6]))

                        # ИМПОРТ ПОТЕРЬ/БРАКА (ждем 2 колонки)
                        elif first_cell == "Тип материала" and len(row) >= 2:
                            val = row[1] if not isinstance(row[1], str) else float(row[1].replace('%', '').replace(',', '.'))
                            cur.execute("INSERT INTO scrap VALUES (%s, %s)", (row[0], val))

                        # ИМПОРТ ПОСТАВЩИКОВ (ждем 5 колонок)
                        elif first_cell == "Наименование поставщика" and len(row) >= 5:
                            cur.execute("INSERT INTO suppliers VALUES (%s, %s, %s, %s, %s)",
                                        (row[0], row[1], row[2], int(row[3]) if row[3] else 0, row[4]))

                        # ИМПОРТ ПРОДУКЦИИ
                        elif first_cell == "Тип продукции" and len(row) >= 5 and "Артикул" in str(sheet.cell(row=1, column=3).value or ""):
                            cur.execute("INSERT INTO products VALUES (%s, %s, %s, %s, %s)",
                                        (str(row[2]), row[1], row[0], float(row[3]), float(row[4])))

                        # ИМПОРТ КОЭФФИЦИЕНТОВ
                        elif first_cell == "Тип продукции" and len(row) >= 2:
                            cur.execute("INSERT INTO coefficients VALUES (%s, %s)", (row[0], float(row[1])))

                except Exception as e:
                    print(f"Ошибка при обработке файла {file}: {e}")

            self.conn.commit()

    def setup_ui(self):
        """Отрисовка главного окна: Логотип -> Навигация -> Контент"""
        try:
            self.logo_img = tk.PhotoImage(file=os.path.join(BASE_PATH, LOGO_FILE))
            tk.Label(self.root, image=self.logo_img, bg=C_MAIN).pack(pady=20)
        except:
            tk.Label(self.root, text="СИСТЕМА МАТЕРИАЛОВ", font=(FONT_NAME, 24, "bold"), fg=C_ACCENT, bg=C_MAIN).pack(pady=20)

        self.main_frame = tk.Frame(self.root, bg=C_MAIN)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Левое меню (Навигация)
        self.nav_bar = tk.Frame(self.main_frame, bg=C_SEC, width=200)
        self.nav_bar.pack(side=tk.LEFT, fill=tk.Y)

        # Кнопки переключения экранов
        menu = [("Склад материалов", self.show_materials), ("Поставщики", self.show_suppliers)]
        for text, cmd in menu:
            btn = tk.Button(self.nav_bar, text=text, command=cmd, bg=C_ACCENT, fg="white",
                            font=(FONT_NAME, F_SIZE), relief=tk.FLAT, pady=10)
            btn.pack(fill=tk.X, padx=10, pady=10)

        # Правая часть окна (Контент)
        self.content = tk.Frame(self.main_frame, bg=C_MAIN)
        self.content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=20)
        self.show_materials() # Стартовый экран

    def clear(self):
        """Удаляет все виджеты с экрана контента перед отрисовкой нового раздела"""
        for w in self.content.winfo_children(): w.destroy()

    def calculate_party_cost(self, stock, min_qty, pack_size, price):
        """
        РАССЧИТЫВАЕМЫЙ АТРИБУТ (Критерий В1Д1):
        Если остаток меньше минимума, вычисляем разницу,
        округляем её до целых упаковок и умножаем на цену.
        """
        if stock >= min_qty: return 0.0
        diff = min_qty - stock
        needed_packs = math.ceil(diff / pack_size)
        total_qty = needed_packs * pack_size
        return round(total_qty * price, 2)

    def show_materials(self):
        """Раздел склада материалов: Отображение в виде карточек (как на макете)"""
        self.clear()
        tk.Label(self.content, text="Список материалов", font=(FONT_NAME, 20), fg=C_ACCENT, bg=C_MAIN).pack(pady=10)

        tk.Button(self.content, text="+ Добавить материал", command=self.open_material_form,
                  bg=C_ACCENT, fg="white", font=(FONT_NAME, 12), pady=5).pack(pady=10)

        # Используем Canvas, так как обычный Frame не умеет скроллиться
        self.canvas = tk.Canvas(self.content, bg=C_MAIN, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.content, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg=C_MAIN)

        self.scrollable_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Запрос: объединяем таблицу материалов с таблицей типов
        query = "SELECT m.id, m.name, t.name, m.min_qty, m.stock, m.price, m.pack_size, m.unit FROM materials m JOIN material_types t ON m.type_id = t.id"
        with self.conn.cursor() as cur:
            cur.execute(query)
            for row in cur.fetchall():
                self.create_material_card(row)

    def create_material_card(self, data):
        """Отрисовка одной карточки материала (согласно макету)"""
        m_id, name, m_type, min_q, stock, price, pack, unit = data
        party_cost = self.calculate_party_cost(stock, min_q, pack, price)

        card = tk.Frame(self.scrollable_frame, bg=C_MAIN, highlightbackground="grey", highlightthickness=1, padx=15, pady=10)
        card.pack(fill=tk.X, padx=20, pady=10)

        top_row = tk.Frame(card, bg=C_MAIN)
        top_row.pack(fill=tk.X)
        tk.Label(top_row, text=f"{m_type} | {name}", font=(FONT_NAME, 16, "bold"), bg=C_MAIN, fg=C_ACCENT).pack(side=tk.LEFT)
        tk.Label(top_row, text=f"Стоимость партии: {party_cost:.2f} р.", font=(FONT_NAME, 16), bg=C_MAIN).pack(side=tk.RIGHT)

        tk.Label(card, text=f"Минимальное количество: {min_q} {unit}", font=(FONT_NAME, 12), bg=C_MAIN).pack(anchor="w")
        tk.Label(card, text=f"Количество на складе: {stock} {unit}", font=(FONT_NAME, 12), bg=C_MAIN).pack(anchor="w")
        tk.Label(card, text=f"Цена: {price:.2f} р / Единица измерения: {unit}", font=(FONT_NAME, 12), bg=C_MAIN).pack(anchor="w")

        btn_row = tk.Frame(card, bg=C_MAIN)
        btn_row.pack(fill=tk.X, pady=5)

        def edit():
            with self.conn.cursor() as cur:
                cur.execute("SELECT name, type_id, price, stock, min_qty, pack_size, unit FROM materials WHERE id=%s", (m_id,))
                self.open_material_form(cur.fetchone())

        tk.Button(btn_row, text="Редактировать", command=edit, bg=C_ACCENT, fg="white", font=(FONT_NAME, 10)).pack(side=tk.LEFT, padx=5)

    def show_suppliers(self):
        """Раздел просмотра списка поставщиков (Табличный вид)"""
        self.clear()
        tk.Label(self.content, text="Список поставщиков материалов", font=(FONT_NAME, 20), fg=C_ACCENT, bg=C_MAIN).pack(pady=10)

        # Определяем колонки для таблицы
        cols = ("Name", "Type", "INN", "Rating", "StartDate")
        cols_text = ["Наименование", "Тип", "ИНН", "Рейтинг", "Дата начала работы"]

        tree = ttk.Treeview(self.content, columns=cols, show="headings")
        for i in range(len(cols)):
            tree.heading(cols[i], text=cols_text[i])
            tree.column(cols[i], width=150)

        with self.conn.cursor() as cur:
            cur.execute("SELECT name, type, inn, rating, start_date FROM suppliers")
            for row in cur.fetchall():
                tree.insert("", tk.END, values=row)

        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    def open_material_form(self, edit_data=None):
        """Форма добавления/правки материала с валидацией (Е1Д4)"""
        win = tk.Toplevel(self.root)
        win.title("Редактор материала")
        win.geometry("450x600")
        win.configure(bg=C_MAIN)

        field_info = [("Наименование", "name"), ("Тип материала", "type"), ("Количество на складе", "stock"),
                      ("Единица измерения", "unit"), ("Количество в упаковке", "pack"), ("Минимальное количество", "min"), ("Цена единицы", "price")]
        entries = {}

        for label_text, key in field_info:
            tk.Label(win, text=label_text, bg=C_MAIN, font=(FONT_NAME, 12)).pack(pady=5)
            if key == "type":
                with self.conn.cursor() as cur:
                    cur.execute("SELECT name FROM material_types")
                    types = [r[0] for r in cur.fetchall()]
                combo = ttk.Combobox(win, values=types, state="readonly")
                combo.pack(pady=5)
                entries[key] = combo
            else:
                entry = tk.Entry(win)
                entry.pack(pady=5)
                entries[key] = entry

        if edit_data:
            with self.conn.cursor() as cur:
                cur.execute("SELECT name FROM material_types WHERE id=%s", (edit_data[1],))
                t_name = cur.fetchone()[0]
            entries['name'].insert(0, edit_data[0]); entries['type'].set(t_name); entries['price'].insert(0, edit_data[2])
            entries['stock'].insert(0, edit_data[3]); entries['min'].insert(0, edit_data[4]); entries['pack'].insert(0, edit_data[5])
            entries['unit'].insert(0, edit_data[6])

        def save():
            try:
                vals = {k: e.get().strip() for k, e in entries.items()}
                if not all(vals.values()): raise ValueError("Заполните все поля!")

                try:
                    price = float(vals['price'].replace(',', '.'))
                    stock = float(vals['stock'].replace(',', '.'))
                    min_q = float(vals['min'].replace(',', '.'))
                    pack = float(vals['pack'].replace(',', '.'))
                except: raise ValueError("Цена, Склад, Мин. кол-во и Упаковка должны быть числами!")

                if price < 0 or stock < 0 or min_q < 0 or pack < 0:
                    raise ValueError("Значения не могут быть отрицательными!")

                with self.conn.cursor() as cur:
                    cur.execute("SELECT id FROM material_types WHERE name = %s", (vals['type'],))
                    t_id = cur.fetchone()[0]
                    if edit_data:
                        cur.execute("UPDATE materials SET name=%s, type_id=%s, price=%s, stock=%s, min_qty=%s, pack_size=%s, unit=%s WHERE name=%s",
                                    (vals['name'], t_id, price, stock, min_q, pack, vals['unit'], edit_data[0]))
                    else:
                        cur.execute("INSERT INTO materials (name, type_id, price, stock, min_qty, pack_size, unit) VALUES (%s,%s,%s,%s,%s,%s,%s)",
                                    (vals['name'], t_id, price, stock, min_q, pack, vals['unit']))
                    self.conn.commit()
                messagebox.showinfo("Успех", "Сохранено!")
                win.destroy()
                self.show_materials()
            except Exception as e:
                messagebox.showerror("Ошибка ввода", str(e))

        tk.Button(win, text="Сохранить", command=save, bg=C_ACCENT, fg="white", font=(FONT_NAME, 12), pady=10).pack(pady=20)

if __name__ == "__main__":
    root = tk.Tk()
    app = MaterialApp(root)
    root.mainloop()